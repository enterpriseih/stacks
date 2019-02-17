from airflow import DAG
from airflow.operators.python_operator import PythonOperator
from airflow.hooks.mysql_hook import MySqlHook
from airflow.contrib.hooks.ftp_hook import FTPHook

import io
import openpyxl
from datetime import datetime, timedelta
import pendulum
local_tz = pendulum.timezone('Asia/Shanghai')

dag = DAG(
    'example_ftp_excel_mysql',
    catchup=False,
    schedule_interval=None,
    default_args={
        'owner': 'airflow',
        'depends_on_past': False,
        'start_date': datetime(2018, 10, 29, tzinfo=local_tz),
        'email': ['airflow@example.com'],
        'email_on_failure': True,
        'email_on_retry': True,
        'retries': 0,
        'retry_delay': timedelta(minutes=1),
    }
)
SHEET = 'Sheet1'
HEADER = ['field1', 'field2']
SOURCE_FIELDS = ['field1', 'field2']
TARGET_FIELDS = ['field3', 'field4']
AUX_FIELDS = ['field5', 'field6', 'field7', 'field8', 'field9']

now = datetime.now(local_tz)


def recursive_list(ftp, path, file_extension):
    result = []
    files = ftp.list_directory(path)
    for file in files:
        if file.endswith(file_extension):
            result.append(path + "/" + file)
        else:
            result = result + \
                recursive_list(ftp, path + "/"+file, file_extension)
    return result


def get_merged_cell_value(ws, row, col):
    value = None
    for range in ws.merged_cells.ranges:
        if row <= range.max_row and row >= range.min_row \
                and col <= range.max_col and col >= range.min_col:
            value = ws["{0}{1}".format(
                openpyxl.utils.get_column_letter(
                    range.min_col), range.min_row)].value
            break
    return value


def main():
    ftp = FTPHook(ftp_conn_id='my_ftp')
    ftp.get_conn().encoding = 'utf-8'

    result = []

    # find all xlsx files
    # files = recursive_list(ftp, '/my_dir/', '.xlsx')

    files = ftp.list_directory('/my_dir/')
    for file in files:
        try:
            buffer = io.BytesIO()
            remote_path = '/my_dir/' + file

            # downloading might take time
            ftp.retrieve_file(remote_path, buffer)

            # open excel sheet
            wb = openpyxl.load_workbook(filename=buffer, data_only=True)
            ws = wb[SHEET]

            for i, row in enumerate(ws.rows):
                # skip header
                if i == 0:
                    continue

                # parse row
                item = {}
                for j, cell in enumerate(row):
                    item[HEADER[j]] = cell.value

                # get field values
                r = [item[SOURCE_FIELDS[k]]
                     for k, field in enumerate(TARGET_FIELDS)]

                r = r + [now, 'airflow', now, 'airflow', remote_path]
                # append
                result.append(r)

        except Exception as e:
            print('error file {0}: {1}'.format(file, str(e)))
    # dump
    target = MySqlHook(mysql_conn_id='my_mysql')
    target.insert_rows(
        table='my_table',
        rows=result,
        commit_every=5000,
        target_fields=TARGET_FIELDS+AUX_FIELDS
    )

    return


t1 = PythonOperator(
    task_id="main",
    python_callable=main,
    queue="<worker2_ip>",
    dag=dag
)
